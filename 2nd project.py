import pandas as pd
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup

headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36"}
requests.packages.urllib3.disable_warnings()
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS += ':HIGH:!DH:!aNULL'

filename = 'NewsResult_20181001-20181231.xlsx'
wb = load_workbook(filename)
ws=wb["sheet"]

for row in range(2,506):
    try:
        srow=str(row)
        ui="R"+srow
        urlinfo=ws[ui]
        url=urlinfo.value
        res=requests.get(url,headers=headers,verify=False)
        soup=BeautifulSoup(res.text,"lxml")
        if 'khan' in url:
            context=soup.find("div",attrs={"class":"art_body"}).get_text()
        if "donga" in url:
            context=soup.find("div",attrs={"class":"article_txt"}).get_text()
        if "hani" in url:
            context=soup.find("div",attrs={"class":"text"}).get_text()
        if "chosun" in url:
            context=soup.find("section",attrs={"class":"article-body"}).get_text()
        if "ihalla" in url:
            context=soup.find("div",attrs={"class":"article_txt"}).get_text()
        ws.cell(row=row,column=27,value=context)
        wb.save('NewsResult_20181001-20181231.xlsx')
    except:
        ws.cell(row=row,column=27,value='추출 불가')
        wb.save('NewsResult_20181001-20181231.xlsx')
